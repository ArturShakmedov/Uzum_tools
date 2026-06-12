"""Парсер Excel «Новые комиссии c калькулятором.xlsx» → таблица uzum_categories.

Лист «Комиссия за продажу (РУС)»: заголовки на 3-й строке (row index 2).
Колонки: category ID, category1_ru…category6_ru, базовые comm FBO %/comm FBS %
(колонки 8/9) и СКИДОЧНЫЕ comm FBO %.1/comm FBS %.1 (колонки 14/15, действуют
до 30.06.2026). Берём скидочную, если она задана, иначе базовую. Комиссии — доли
(0.1 = 10%).

Запуск:
    python scripts/parse_commissions.py ["путь к .xlsx"]

Реализовано на openpyxl (уже зависимость, без тяжёлого pandas); логика идентична
pandas header=2 с дедупликацией дублей имён колонок в *.1.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy import delete

from database.connection import init_db, session_scope
from database.models import UzumCategory
from utils.logger import get_logger

log = get_logger(__name__)

_SHEET = "Комиссия за продажу (РУС)"
_HEADER_ROW = 3  # 1-индекс (pandas header=2)
_CATEGORY_COLS = [f"category{i}_ru" for i in range(1, 7)]

# Крупногабарит (КГТ) → логистический сбор 20000. Ключевые слова в категориях.
_KGT_KEYWORDS = (
    "крупная бытовая техника", "телевизор", "холодильник", "морозильн",
    "стиральн", "посудомоечн", "кондиционер", "духов", "плита", "варочн",
    "мебель", "диван", "шкаф", "кровать", "матрас", "велосипед",
    "беговая дорожка", "sim", "сим-карт",
)


def _header_indexes(header: list) -> dict:
    """Сопоставить нужные имена колонок их индексам.

    Для дублей (comm FBO %, comm FBS %) первая встреча — базовая, вторая — *.1.
    """
    def find(name: str, occurrence: int = 0) -> int | None:
        seen = 0
        for i, h in enumerate(header):
            if (h or "").strip() == name:
                if seen == occurrence:
                    return i
                seen += 1
        return None

    idx = {
        "id": find("category ID"),
        "comm_fbo_base": find("comm FBO %", 0),
        "comm_fbs_base": find("comm FBS %", 0),
        "comm_fbo_disc": find("comm FBO %", 1),  # = comm FBO %.1
        "comm_fbs_disc": find("comm FBS %", 1),  # = comm FBS %.1
    }
    for c in _CATEGORY_COLS:
        idx[c] = find(c)
    return idx


def _num(value) -> float | None:
    try:
        return float(value) if value is not None and value != "" else None
    except (TypeError, ValueError):
        return None


def parse_categories(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[_SHEET]

    header = [c for c in next(ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))]
    idx = _header_indexes(header)
    if idx["id"] is None:
        raise ValueError("Не найдена колонка 'category ID' — проверьте лист/строку заголовка.")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=_HEADER_ROW + 1, values_only=True):
        cat_id = row[idx["id"]]
        if cat_id is None:
            continue
        cats = [str(row[idx[c]]).strip() for c in _CATEGORY_COLS
                if idx[c] is not None and row[idx[c]] not in (None, "")]
        if not cats:
            continue
        search_text = " ".join(cats).lower()
        # Скидочная комиссия в приоритете, иначе базовая.
        comm_fbo = _num(row[idx["comm_fbo_disc"]]) if idx["comm_fbo_disc"] is not None else None
        if comm_fbo is None:
            comm_fbo = _num(row[idx["comm_fbo_base"]]) or 0.0
        comm_fbs = _num(row[idx["comm_fbs_disc"]]) if idx["comm_fbs_disc"] is not None else None
        if comm_fbs is None:
            comm_fbs = _num(row[idx["comm_fbs_base"]]) or 0.0

        rows.append({
            "id": int(cat_id),
            "display_name": " -> ".join(cats),
            "search_text": search_text,
            "comm_fbo": float(comm_fbo),
            "comm_fbs": float(comm_fbs),
            "is_kgt": any(kw in search_text for kw in _KGT_KEYWORDS),
        })
    return rows


def main(xlsx_path: str) -> int:
    path = Path(xlsx_path)
    if not path.exists():
        log.error("Файл не найден: %s", path)
        return 1

    init_db()
    rows = parse_categories(path)
    log.info("Распознано категорий: %d", len(rows))

    with session_scope() as session:
        session.execute(delete(UzumCategory))  # перезаливаем справочник целиком
        session.bulk_insert_mappings(UzumCategory, rows)

    kgt = sum(1 for r in rows if r["is_kgt"])
    log.info("Записано в uzum_categories: %d (из них КГТ: %d)", len(rows), kgt)

    # Контроль: показываем, что реально легло в БД (search_text в нижнем регистре).
    from sqlalchemy import func, select  # noqa: PLC0415
    with session_scope() as session:
        total = session.execute(select(func.count()).select_from(UzumCategory)).scalar_one()
        sample = session.execute(
            select(UzumCategory.id, UzumCategory.search_text).limit(3)
        ).all()
    log.info("В БД сейчас категорий: %d. Первые 3 search_text:", total)
    for cid, st in sample:
        log.info("  id=%s search_text=%r", cid, st)
    print(f"OK: {total} категорий в uzum_categories (КГТ: {kgt}).")
    return 0


if __name__ == "__main__":
    default = "/run/media/artur/33e890b4-e9ce-4308-8bb8-c1559b8054c5/down/Новые комиссии c калькулятором.xlsx"
    raise SystemExit(main(sys.argv[1] if len(sys.argv) > 1 else default))
