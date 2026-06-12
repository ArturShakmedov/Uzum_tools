"""Генерация Excel-отчёта по невозвратам (openpyxl)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.analytics import NO_NAME, LossRow

_HEADERS = [
    "ID Заказа", "Дата", "SKU", "Штрихкод", "Название", "Статус", "Дней", "Итог",
]
# Ширина колонок под содержимое.
_WIDTHS = [12, 12, 26, 16, 32, 14, 7, 30]
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=12, color="1F4E78")
# Подсветка строки по вердикту.
_FILL_LOST = PatternFill("solid", fgColor="F8CBAD")     # ⚠️ не вернули
_FILL_TRANSIT = PatternFill("solid", fgColor="FFF2CC")  # 🚚 в транзите

_DEFAULT_TITLE = "Отчёт по невозвратам"


def build_losses_xlsx(
    rows: list[LossRow], path: str | Path, *, title: str = _DEFAULT_TITLE
) -> Path:
    """Сформировать .xlsx с отчётом по невозвратам и вернуть путь к файлу.

    title — заголовок таблицы внутри файла (строка 1, объединённая по колонкам).
    """
    path = Path(path)
    ncols = len(_HEADERS)
    wb = Workbook()
    ws = wb.active
    ws.title = "Невозвраты"

    # Строка 1 — заголовок отчёта (на всю ширину таблицы).
    ws.append([title] + [""] * (ncols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tcell = ws.cell(row=1, column=1)
    tcell.font = _TITLE_FONT
    tcell.alignment = Alignment(horizontal="left", vertical="center")

    # Строка 2 — шапка таблицы.
    ws.append(_HEADERS)
    for col, width in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
        cell = ws.cell(row=2, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A3"

    for r in rows:
        sku = str(r.sku_id) if r.sku_id is not None else (r.article or "—")
        ws.append([
            r.order_id,
            r.event_date.strftime("%Y-%m-%d") if r.event_date else "—",
            sku,
            r.barcode,
            r.title or NO_NAME,  # защитный фолбэк названия
            r.uzum_status,
            r.days_elapsed if r.days_elapsed is not None else "—",
            r.result,
        ])
        fill = _FILL_LOST if "НЕ ВЕРНУЛИ" in r.result else (
            _FILL_TRANSIT if "транзите" in r.result else None
        )
        if fill:
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()  # корректно освобождаем книгу openpyxl
    return path


__all__ = ["build_losses_xlsx"]
